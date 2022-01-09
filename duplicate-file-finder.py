import glob
import hashlib
import os
from openpyxl import Workbook
import time

wb = Workbook()
ws = wb.active

fileHashes = dict()


def md5(fname):
    hash_sha1 = hashlib.sha1()
    hash_md5 = hashlib.md5()
    try:
        with open(fname, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha1.update(chunk)
    except:
        return str(time.time())  # in case the file is being used for example
    data = hash_sha1.hexdigest()
    try:
        with open(fname, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
    except:
        return str(time.time())  # in case the file is being used for example
    return data + hash_md5.hexdigest()


resultsFilename = 'duplicate results.xslx'
if os.path.exists(resultsFilename):
    os.remove(resultsFilename)

duplicateCount = 0
numberOfFiles = 0
index = 1
for filename in glob.iglob('D:\\**',
                           recursive=True):
    # try:
    if os.path.isfile(filename):
        checksum = md5(filename)
        # print(checksum)
        if checksum in fileHashes.keys():
            # duplicate file has been found
            size = os.path.getsize(fileHashes[checksum]) / 1024 / 1024
            print(str(size))
            if size > 1:
                print('Duplicate file has been found')
                openFile = input("Found a duplicate file. Do you wish to open the file %s? [y/n]" % filename)
                if openFile == 'y':
                    os.startfile(filename)
                    os.startfile(fileHashes[checksum])
            ws['A' + str(index)] = filename
            ws['B' + str(index)] = fileHashes[checksum]
            ws['C' + str(index)] = str(size)
            ws['D' + str(index)] = 'MB'
            index += 1
            print("%s is duplicate of %s" % (filename, fileHashes[checksum]))
            duplicateCount += 1
            if duplicateCount % 10 == 0:
                print("There are %s duplicates in the folder so far" % duplicateCount)
            else:
                fileHashes[checksum] = filename
            # print("%s %s" % (filename, md5(filename)))
            # except Exception:
            #     pass
            numberOfFiles += 1
            if numberOfFiles % 100 == 0:
                print("Hashed %s files so far" % numberOfFiles)

            wb.save(resultsFilename)
